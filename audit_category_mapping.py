"""
Phase A-3: Amazon 类目 → Walmart ProductType 映射

导入用户手工 + LLM 构建的映射表 Amazon_to_Walmart_Category_Mapping_v9.xlsx
(7008 条 Walmart Product Type, 含推荐 Amazon 路径 / 认证 / IP 风险 / 合规说明)

提供:
- load_mapping_xlsx(xlsx_path)  — 一次性全量导入
- predict_walmart_ptype(amazon_path, amazon_l1, amazon_bsr_id, conn)
    返回 (suggested_ptype, confidence, cert_required, ip_risk, compliance_notes)
"""
import argparse
from pathlib import Path

import openpyxl
import psycopg2
from psycopg2.extras import execute_values


DB_CONN = "dbname=uspto user=nextderboy"


CONF_MAP = {
    "LLM高": "0.95",
    "高":    "0.90",
    "中":    "0.70",
    "低":    "0.50",
    "LLM中": "0.75",
    "LLM低": "0.55",
}


def _clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s and s.upper() != "NONE" else None


def _conf(v):
    if not v:
        return None
    s = str(v).strip()
    # 文字 → 数字
    if s in CONF_MAP:
        return CONF_MAP[s]
    # 已经是小数
    try:
        return str(float(s))
    except ValueError:
        return None


def load_mapping_xlsx(xlsx_path):
    """读取 '完整映射表' sheet → amazon_walmart_category_map"""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if "完整映射表" not in wb.sheetnames:
        raise ValueError(f"xlsx 中未找到 '完整映射表' sheet: {wb.sheetnames}")
    ws = wb["完整映射表"]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h).strip() if h else "" for h in rows[0]]
    data = rows[1:]

    col = {name: idx for idx, name in enumerate(header)}

    parsed = []
    for r in data:
        walmart_category = _clean(r[col.get("Walmart Category")])
        walmart_ptg      = _clean(r[col.get("Walmart PTG (Product Type Group)")])
        walmart_pt       = _clean(r[col.get("Walmart Product Type")])
        if not walmart_pt:
            continue
        rec_l1  = _clean(r[col.get("推荐Amazon L1")])
        rec_l2  = _clean(r[col.get("推荐Amazon L2")])
        rec_l3  = _clean(r[col.get("推荐Amazon L3")])
        rec_bsr = _clean(r[col.get("推荐Amazon BSR类目")])
        rec_path = _clean(r[col.get("推荐Amazon路径")])
        conf_raw = _clean(r[col.get("映射置信度")])
        conf = _conf(conf_raw)
        src = _clean(r[col.get("映射来源")])
        sample = r[col.get("样本数")] or 0
        cert   = _clean(r[col.get("所需认证/文件")])
        ip_risk = _clean(r[col.get("IP侵权风险")])
        compliance_notes = _clean(r[col.get("合规说明")])
        sales_qual = _clean(r[col.get("销售资质要求")])
        remark = _clean(r[col.get("备注")])

        # 构造 amazon_path (L1>L2>L3) 用于匹配
        if not rec_path:
            parts = [p for p in (rec_l1, rec_l2, rec_l3) if p]
            rec_path = " > ".join(parts) if parts else None

        parsed.append({
            "amazon_path": rec_path,
            "walmart_product_type": walmart_pt,
            "walmart_category": walmart_category,
            "walmart_ptg": walmart_ptg,
            "recommended_amazon_l1": rec_l1,
            "recommended_amazon_l2": rec_l2,
            "recommended_amazon_l3": rec_l3,
            "recommended_amazon_bsr": rec_bsr,
            "recommended_amazon_path": rec_path,
            "confidence": conf,
            "history_sample_count": int(sample) if sample else 0,
            "source": src,
            "cert_required_text": cert,
            "ip_risk_level": ip_risk,
            "compliance_notes": compliance_notes,
            "sales_qualification": sales_qual,
            "notes": remark,
            "is_safe": (cert is None or cert in ("无特殊限制", "无")) and ip_risk != "高",
        })

    print(f"解析 {len(parsed)} 条映射")

    conn = psycopg2.connect(DB_CONN)
    cur = conn.cursor()
    # TRUNCATE 重建 (用户映射是权威)
    cur.execute("TRUNCATE amazon_walmart_category_map RESTART IDENTITY")

    COLS = [
        "amazon_path", "walmart_product_type",
        "walmart_category", "walmart_ptg",
        "recommended_amazon_l1", "recommended_amazon_l2", "recommended_amazon_l3",
        "recommended_amazon_bsr", "recommended_amazon_path",
        "confidence", "history_sample_count", "source",
        "cert_required_text", "ip_risk_level",
        "compliance_notes", "sales_qualification", "notes",
        "is_safe",
    ]
    values = []
    seen = set()
    for p in parsed:
        key = (p["amazon_path"], p["walmart_product_type"])
        if key in seen:
            continue
        seen.add(key)
        values.append(tuple(p[c] for c in COLS))

    execute_values(cur,
        f"INSERT INTO amazon_walmart_category_map ({','.join(COLS)}) VALUES %s",
        values, page_size=500
    )
    conn.commit()
    print(f"入库 {len(values)} 条 (去重后)")

    # 统计
    cur.execute("""
        SELECT source, COUNT(*) FROM amazon_walmart_category_map
        GROUP BY source ORDER BY COUNT(*) DESC
    """)
    for src, n in cur.fetchall():
        print(f"  {src}: {n}")
    cur.close()
    conn.close()
    return len(values)


SELECT_COLS = """
    walmart_product_type, walmart_category, confidence, cert_required_text,
    ip_risk_level, compliance_notes, is_safe, source
"""


def _row_to_cand(r, match_level):
    return {
        "walmart_product_type": r[0], "walmart_category": r[1],
        "confidence": float(r[2]) if r[2] else None,
        "cert_required": r[3], "ip_risk": r[4],
        "compliance_notes": r[5], "is_safe": r[6],
        "source": r[7], "match_level": match_level,
    }


def predict_walmart_ptype(conn, amazon_path=None, **kw):
    """
    给定 Amazon category_path 返回最匹配的 Walmart ProductType 候选

    Amazon path 格式: "L1 > L2 > L3 > BSR" (4 段) 或少于 4 段
    匹配优先级 (严格 → 宽松):
      1. 完整路径与 recommended_amazon_path 相等
      2. BSR (末段) 匹配 recommended_amazon_bsr
      3. L3 (倒数第二段, 若有 ≥4 段) 匹配 recommended_amazon_l3
      4. L3 精确匹配末段 (若只有 3 段)
    不做 L1 粗匹配 (噪声过大)
    """
    cur = conn.cursor()
    candidates = []
    seen_pt = set()

    def add_rows(rows, level):
        for r in rows:
            if r[0] in seen_pt:
                continue
            seen_pt.add(r[0])
            candidates.append(_row_to_cand(r, level))

    # 1. 完整路径匹配
    if amazon_path:
        cur.execute(
            f"SELECT {SELECT_COLS} FROM amazon_walmart_category_map "
            f"WHERE recommended_amazon_path = %s "
            f"ORDER BY confidence DESC NULLS LAST LIMIT 5",
            (amazon_path,))
        add_rows(cur.fetchall(), "exact_path")

    # 2. BSR (末段) 匹配
    parts = [p.strip() for p in (amazon_path or "").split(" > ") if p.strip()]
    bsr = parts[-1] if parts else None
    if bsr and len(candidates) < 3:
        cur.execute(
            f"SELECT {SELECT_COLS} FROM amazon_walmart_category_map "
            f"WHERE recommended_amazon_bsr = %s "
            f"ORDER BY confidence DESC NULLS LAST LIMIT 5",
            (bsr,))
        add_rows(cur.fetchall(), "bsr")

    # 3. L3 (倒数第二段) 匹配 - 仅在 4 段或更多时
    if len(parts) >= 4 and len(candidates) < 3:
        l3 = parts[-2]
        cur.execute(
            f"SELECT {SELECT_COLS} FROM amazon_walmart_category_map "
            f"WHERE recommended_amazon_l3 = %s "
            f"ORDER BY confidence DESC NULLS LAST LIMIT 5",
            (l3,))
        add_rows(cur.fetchall(), "l3")

    # 4. 末段作为 L3 匹配 (3 段结构)
    if len(parts) == 3 and len(candidates) < 3:
        cur.execute(
            f"SELECT {SELECT_COLS} FROM amazon_walmart_category_map "
            f"WHERE recommended_amazon_l3 = %s "
            f"ORDER BY confidence DESC NULLS LAST LIMIT 5",
            (parts[-1],))
        add_rows(cur.fetchall(), "l3_from_end")

    cur.close()
    return candidates[:5]


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--xlsx", default="/Users/nextderboy/Downloads/Amazon_to_Walmart_Category_Mapping_v9.xlsx")
    p.add_argument("--test-asin", help="测试: 给个 ASIN 看预测结果")
    args = p.parse_args()

    if args.test_asin:
        conn = psycopg2.connect(DB_CONN)
        cur = conn.cursor()
        cur.execute("""
            SELECT category_path, root_category_id
            FROM products_stage WHERE asin = %s LIMIT 1
        """, (args.test_asin,))
        row = cur.fetchone()
        if not row:
            print(f"未找到 ASIN {args.test_asin}")
            return
        cat_path, root_id = row
        print(f"Amazon 类目: {cat_path}")
        print(f"Root ID: {root_id}")
        # 从 path 抽 L3/BSR
        parts = cat_path.split(" > ") if cat_path else []
        bsr = parts[-1] if parts else None
        cands = predict_walmart_ptype(conn, amazon_path=cat_path, amazon_bsr=bsr)
        for i, c in enumerate(cands[:5]):
            print(f"\n候选 {i+1} ({c['match_level']}):")
            print(f"  Walmart PT: {c['walmart_product_type']}")
            print(f"  类别: {c['walmart_category']}")
            print(f"  置信度: {c['confidence']}")
            print(f"  认证: {c['cert_required']}")
            print(f"  IP 风险: {c['ip_risk']}")
            print(f"  合规: {c['compliance_notes']}")
            print(f"  是否安全(搬运): {c['is_safe']}")
        cur.close()
        conn.close()
        return

    # 默认: 导入映射表
    load_mapping_xlsx(args.xlsx)


if __name__ == "__main__":
    main()

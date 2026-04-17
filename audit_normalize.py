"""
Amazon 采集 xlsx → products_stage 表

采集文件 39 列 (按位置映射), 典型文件如 batch_20260417_032822.xlsx 的 '采集结果' sheet

用法:
    python3 audit_normalize.py /path/to/batch.xlsx
    python3 audit_normalize.py /path/to/dir/*.xlsx
"""
import argparse
import glob
import json
import os
import re
import sys
import time
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl
import psycopg2
from psycopg2.extras import Json, execute_values


DB_CONN = "dbname=uspto user=nextderboy"

# 采集 xlsx 列名 → products_stage 字段 (按名称精确匹配)
COLUMN_MAP = {
    "ASIN (商品ID)": "asin",
    "商品标题": "title",
    "品牌": "brand",
    "商品类型": "amazon_product_type",    # 注意: Amazon 商品类型, 不是沃尔玛 ProductType
    "制造商": "manufacturer",
    "产品型号": "model_number",
    "部件编号": "part_number",
    "原产国": "origin_country",
    "畅销排名": "bsr",
    "当前价格": "current_price_raw",
    "BuyBox 价格": "buybox_price_raw",
    "是否 FBA 发货": "fba_raw",
    "商品图片链接": "image_urls",
    "五点描述": "bullet_points",
    "长描述": "long_description",
    "UPC 列表": "upc",
    "EAN 列表": "ean",
    "根类目 ID": "root_category_id",
    "类目 ID 链": "category_chain_ids",
    "类目路径树": "category_path",
    "商品采集时间": "collected_at_raw",
}

PRICE_RE = re.compile(r"[\d,]+\.?\d*")


def _parse_price(s):
    """'$39.99' / 'N/A' / '' → Decimal or None"""
    if s is None:
        return None
    s = str(s).strip()
    if not s or s.upper() in ("N/A", "NONE", "NULL"):
        return None
    m = PRICE_RE.search(s.replace(",", ""))
    if not m:
        return None
    try:
        return Decimal(m.group(0))
    except InvalidOperation:
        return None


def _parse_fba(s):
    if s is None:
        return None
    s = str(s).strip().upper()
    if s == "FBA":
        return True
    if s in ("FBM", "SELF", "MERCHANT"):
        return False
    return None


def _parse_collected_at(s):
    if not s:
        return None
    s = str(s).strip()
    # 格式 '2026-04-17 11:30:03'
    try:
        from datetime import datetime
        return datetime.strptime(s, "%Y-%m-%d %H:%M:%S")
    except ValueError:
        return None


def _clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s and s.upper() != "N/A" else None


def normalize_xlsx(xlsx_path, batch_file_name=None):
    """
    读取 xlsx, 返回 list of dicts 准备写入 products_stage
    """
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        raise FileNotFoundError(xlsx_path)

    batch_file = batch_file_name or xlsx_path.name

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if "采集结果" in wb.sheetnames:
        ws = wb["采集结果"]
    else:
        ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        return []
    header = [str(h).strip() if h is not None else "" for h in header]

    # header name → index
    header_idx = {h: i for i, h in enumerate(header)}

    # 构建实际使用的列映射
    used_map = {}  # field_name → column_index
    for col_name, field in COLUMN_MAP.items():
        idx = header_idx.get(col_name)
        if idx is not None:
            used_map[field] = idx

    missing = [k for k in ("asin", "title") if k not in used_map]
    if missing:
        raise ValueError(f"采集文件缺少必要列: {missing}. 实际表头: {header[:5]}...")

    products = []
    for row in rows_iter:
        if not row or not row[used_map["asin"]]:
            continue
        d = {}
        for field, idx in used_map.items():
            d[field] = row[idx] if idx < len(row) else None

        asin = _clean(d.get("asin"))
        if not asin:
            continue

        # 派生字段
        price = _parse_price(d.get("current_price_raw"))
        if price is None:
            price = _parse_price(d.get("buybox_price_raw"))

        record = {
            "batch_file": batch_file,
            "asin": asin,
            "title": _clean(d.get("title")),
            "brand": _clean(d.get("brand")),
            "product_type": _clean(d.get("amazon_product_type")),
            "manufacturer": _clean(d.get("manufacturer")),
            "model_number": _clean(d.get("model_number")),
            "part_number": _clean(d.get("part_number")),
            "origin_country": _clean(d.get("origin_country")),
            "bullet_points": _clean(d.get("bullet_points")),
            "long_description": _clean(d.get("long_description")),
            "image_urls": _clean(d.get("image_urls")),
            "upc": _clean(d.get("upc")),
            "ean": _clean(d.get("ean")),
            "root_category_id": _clean(d.get("root_category_id")),
            "category_chain_ids": _clean(d.get("category_chain_ids")),
            "category_path": _clean(d.get("category_path")),
            "price": price,
            "bsr": _clean(d.get("bsr")),
            "is_fba": _parse_fba(d.get("fba_raw")),
            "collected_at": _parse_collected_at(d.get("collected_at_raw")),
            "raw_data": {k: (str(v) if v is not None else None) for k, v in d.items()},
        }
        products.append(record)

    return products


INSERT_SQL = """
INSERT INTO products_stage (
    batch_file, asin, title, brand, product_type, manufacturer,
    model_number, part_number, origin_country,
    bullet_points, long_description, image_urls,
    upc, ean, root_category_id, category_chain_ids, category_path,
    price, bsr, is_fba, raw_data, collected_at
) VALUES %s
ON CONFLICT (batch_file, asin) DO UPDATE SET
    title = EXCLUDED.title,
    brand = EXCLUDED.brand,
    product_type = EXCLUDED.product_type,
    bullet_points = EXCLUDED.bullet_points,
    long_description = EXCLUDED.long_description,
    image_urls = EXCLUDED.image_urls,
    category_path = EXCLUDED.category_path,
    price = EXCLUDED.price,
    raw_data = EXCLUDED.raw_data,
    imported_at = NOW()
"""

COLS_ORDER = [
    "batch_file", "asin", "title", "brand", "product_type", "manufacturer",
    "model_number", "part_number", "origin_country",
    "bullet_points", "long_description", "image_urls",
    "upc", "ean", "root_category_id", "category_chain_ids", "category_path",
    "price", "bsr", "is_fba", "raw_data", "collected_at",
]


def import_xlsx(xlsx_path, verbose=True):
    t0 = time.time()
    products = normalize_xlsx(xlsx_path)
    t_read = time.time() - t0
    if verbose:
        print(f"  解析: {len(products)} 产品, 耗时 {t_read:.1f}s")
    if not products:
        return 0

    conn = psycopg2.connect(DB_CONN)
    cur = conn.cursor()
    batch_size = 500
    inserted = 0
    t_ins = time.time()
    for i in range(0, len(products), batch_size):
        batch = products[i:i + batch_size]
        values = []
        for p in batch:
            row = tuple(
                Json(p[c]) if c == "raw_data" else p[c]
                for c in COLS_ORDER
            )
            values.append(row)
        execute_values(cur, INSERT_SQL, values, page_size=200)
        inserted += len(batch)
    conn.commit()
    cur.close()
    conn.close()
    if verbose:
        print(f"  DB 写入: {inserted} 行, 耗时 {time.time() - t_ins:.1f}s")
    return inserted


def main():
    p = argparse.ArgumentParser()
    p.add_argument("paths", nargs="+", help="xlsx 文件或 glob 模式")
    args = p.parse_args()

    files = []
    for pth in args.paths:
        if "*" in pth or "?" in pth:
            files.extend(sorted(glob.glob(pth)))
        elif os.path.isdir(pth):
            files.extend(sorted(Path(pth).glob("*.xlsx")))
        else:
            files.append(pth)

    if not files:
        print("未找到 xlsx 文件")
        sys.exit(1)

    total = 0
    for f in files:
        print(f"\n=== 导入 {f} ===")
        total += import_xlsx(f)
    print(f"\n=== 合计 {total} 产品 ===")


if __name__ == "__main__":
    main()

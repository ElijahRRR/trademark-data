#!/usr/bin/env python3
"""
TRO案件 → USPTO商标数据库 批量匹配处理
Step 0 ~ Step 5 完整流程
"""

import re
import json
import psycopg2
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

DB_CONN = "dbname=uspto user=nextderboy"
XLSX_INPUT = "/Users/nextderboy/Projects/tro-scraper-matrix/2026-04-07_merged_export.xlsx"
XLSX_OUTPUT = "/Users/nextderboy/Projects/商标数据/company_profiles_output.xlsx"


def validate_match(query_name, candidate_name):
    """验证匹配结果的相关性 (0-1)"""
    q_words = set(w.lower().strip('.,;') for w in query_name.split() if len(w) > 1)
    c_words = set(w.lower().strip('.,;') for w in candidate_name.split() if len(w) > 1)
    stop_words = {'the', 'a', 'an', 'of', 'and', 'for', 'on', 'inc', 'inc.', 'llc',
                  'ltd', 'ltd.', 'co', 'co.', 'corp', 'corporation', 'company', 'limited'}
    q_meaningful = q_words - stop_words
    if not q_meaningful:
        return 0.0
    return len(q_meaningful & (c_words - stop_words)) / len(q_meaningful)


# =====================================================
# Step 0: 数据清洗
# =====================================================

def clean_plaintiff(raw):
    """从原始原告字段提取干净的公司/个人名"""
    if not raw:
        return ''
    s = raw.strip()
    # 去掉 | 后面的描述
    s = s.split('|')[0].strip()
    # 去掉 ; 后面的描述 (如 ";a Delaware corporation")
    # 但保留 "A; B" 这种多原告格式 — 只在描述性短语前截断
    desc_patterns = [
        r';?\s*a [A-Z][a-z]+ (corporation|company|limited|LLC)',
        r';?\s*an? (individual|entity)',
        r';?\s*doing business as',
        r';?\s*Plaintiff is',
        r';?\s*\[Terminated:',
        r';?\s*Esq\.',
        r';\s*\d{3,5}\s+\w+',  # 地址如 "; 3200 Corporate Center Drive"
    ]
    for pat in desc_patterns:
        s = re.split(pat, s, flags=re.IGNORECASE)[0].strip()
    # 去掉尾部多余标点
    s = s.rstrip(';,| ')
    return s


def clean_brand(raw):
    """从原始品牌字段提取英文品牌名"""
    if not raw:
        return ''
    s = raw.strip()
    # 纯中文 → 保留原样（后续走模糊匹配）
    if not re.search(r'[a-zA-Z0-9]', s):
        return s
    # 中英混合：提取英文部分
    # 模式: "DESIGNICE珠宝" "MILWAUKEE TOOL 密尔沃基工具" "CHRYSLER 克莱斯勒汽车"
    # 策略: 取第一段连续的英文+数字+空格+标点
    # 从头开始匹配英文字符序列
    match = re.match(r"([A-Za-z0-9][A-Za-z0-9\s&'.\-/+!@#$%^*(),:]+)", s)
    if match:
        extracted = match.group(1).strip()
        # 去掉尾部标点
        extracted = extracted.rstrip(' .,;:-')
        if len(extracted) >= 2:
            return extracted.upper()
    # fallback: 去掉所有中文字符
    en_only = re.sub(r'[\u4e00-\u9fff\u3000-\u303f\uff00-\uffef]+', ' ', s).strip()
    en_only = re.sub(r'\s+', ' ', en_only).strip()
    return en_only.upper() if en_only else s


def step0_load_and_clean():
    """读取xlsx，清洗数据，写入tro_cases表"""
    print("=" * 60)
    print("Step 0: 数据清洗")
    print("=" * 60)

    conn = psycopg2.connect(DB_CONN)
    cur = conn.cursor()
    cur.execute("TRUNCATE tro_cases")

    wb = openpyxl.load_workbook(XLSX_INPUT, read_only=True)
    ws = wb.active

    inserted = 0
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        case_number = str(row[1] or '').strip()
        if not case_number:
            skipped += 1
            continue

        date_filed = row[0]
        nature = str(row[3] or '').strip()
        plaintiff_raw = str(row[4] or '').strip()
        brand_raw = str(row[6] or '').strip()

        plaintiff_clean = clean_plaintiff(plaintiff_raw)
        brand_clean = clean_brand(brand_raw)
        brand_eq = (brand_raw == plaintiff_raw) or (brand_clean == plaintiff_clean)

        try:
            cur.execute("""
                INSERT INTO tro_cases (case_number, date_filed, nature_of_suit,
                    plaintiff_raw, brand_raw, plaintiff_clean, brand_clean, brand_eq_plaintiff)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (case_number) DO NOTHING
            """, (case_number, date_filed, nature, plaintiff_raw, brand_raw,
                  plaintiff_clean, brand_clean, brand_eq))
            inserted += 1
        except Exception as e:
            print(f"  跳过 {case_number}: {e}")
            conn.rollback()
            skipped += 1

    conn.commit()
    wb.close()

    # 统计
    cur.execute("SELECT COUNT(*) FROM tro_cases")
    total = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM tro_cases WHERE brand_eq_plaintiff")
    eq_count = cur.fetchone()[0]

    print(f"  写入: {total} 条")
    print(f"  跳过: {skipped} 条")
    print(f"  品牌=原告: {eq_count} 条")

    cur.close()
    conn.close()
    return total


# =====================================================
# Step 1: 去重
# =====================================================

def step1_dedup():
    print("\n" + "=" * 60)
    print("Step 1: 去重")
    print("=" * 60)

    conn = psycopg2.connect(DB_CONN)
    cur = conn.cursor()

    cur.execute("TRUNCATE unique_plaintiffs RESTART IDENTITY")
    cur.execute("TRUNCATE unique_brands RESTART IDENTITY")

    cur.execute("""
        INSERT INTO unique_plaintiffs (plaintiff_clean, case_count)
        SELECT plaintiff_clean, COUNT(*)
        FROM tro_cases
        WHERE plaintiff_clean != ''
        GROUP BY plaintiff_clean
    """)

    cur.execute("""
        INSERT INTO unique_brands (brand_clean, case_count)
        SELECT brand_clean, COUNT(*)
        FROM tro_cases
        WHERE brand_clean != '' AND NOT brand_eq_plaintiff
        GROUP BY brand_clean
    """)

    conn.commit()

    cur.execute("SELECT COUNT(*) FROM unique_plaintiffs")
    p_count = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM unique_brands")
    b_count = cur.fetchone()[0]

    print(f"  唯一原告: {p_count}")
    print(f"  唯一品牌(排除=原告): {b_count}")

    cur.close()
    conn.close()
    return p_count, b_count


# =====================================================
# Step 2: 批量查询
# =====================================================

def step2_batch_query():
    """
    分两轮执行:
      第1轮: 全部精确匹配 (ILIKE, 走GIN索引, 快)
      第2轮: 只对miss的做模糊匹配 (similarity, 慢, 量少)
    """
    print("\n" + "=" * 60)
    print("Step 2: 批量查询")
    print("=" * 60)

    conn = psycopg2.connect(DB_CONN)
    cur = conn.cursor()

    cur.execute("TRUNCATE path_a_results RESTART IDENTITY")
    cur.execute("TRUNCATE path_b_results RESTART IDENTITY")

    # ---- 路径A: 公司名查询 ----
    cur.execute("SELECT id, plaintiff_clean FROM unique_plaintiffs ORDER BY id")
    plaintiffs = cur.fetchall()

    # 第1轮: 精确匹配
    print(f"  路径A 第1轮(精确): {len(plaintiffs)} 个原告...")
    a_exact = 0
    a_miss_ids = []

    for i, (pid, plaintiff) in enumerate(plaintiffs):
        if i % 500 == 0 and i > 0:
            print(f"    进度: {i}/{len(plaintiffs)}")
            conn.commit()

        cur.execute("""
            SELECT DISTINCT t.serial_number, t.mark_identification, o.party_name,
                   t.status_code, m.live_dead
            FROM trademarks t
            JOIN trademark_owners o ON t.serial_number = o.serial_number
            JOIN status_code_mapping m ON t.status_code = m.status_code
            WHERE o.party_name ILIKE %s
              AND m.live_dead = 'LIVE'
        """, (f'%{plaintiff}%',))
        rows = cur.fetchall()

        if rows:
            a_exact += 1
            for row in rows:
                cur.execute("""
                    INSERT INTO path_a_results
                        (plaintiff_clean, match_type, serial_number, mark_identification,
                         owner_name, status_code, live_dead)
                    VALUES (%s, 'exact', %s, %s, %s, %s, %s)
                """, (plaintiff, row[0], row[1], row[2], row[3], row[4]))
            cur.execute("UPDATE unique_plaintiffs SET query_status = 'exact' WHERE id = %s", (pid,))
        else:
            a_miss_ids.append((pid, plaintiff))

    conn.commit()
    print(f"  路径A 第1轮完成: exact={a_exact}, 待模糊={len(a_miss_ids)}")

    # 第2轮: 模糊匹配 (只处理miss的)
    a_fuzzy = 0
    a_miss = 0
    if a_miss_ids:
        print(f"  路径A 第2轮(模糊): {len(a_miss_ids)} 个...")
        for i, (pid, plaintiff) in enumerate(a_miss_ids):
            if i % 100 == 0 and i > 0:
                print(f"    进度: {i}/{len(a_miss_ids)}")
                conn.commit()

            # 用 % 通配 + LIMIT 加速，而非 similarity 全表扫描
            # 取原告名的核心词（前两个词）
            words = plaintiff.split()
            if len(words) >= 2:
                pattern = f'%{words[0]}%{words[1]}%'
            else:
                pattern = f'%{plaintiff}%'

            cur.execute("""
                SELECT DISTINCT t.serial_number, t.mark_identification, o.party_name,
                       t.status_code, m.live_dead
                FROM trademarks t
                JOIN trademark_owners o ON t.serial_number = o.serial_number
                JOIN status_code_mapping m ON t.status_code = m.status_code
                WHERE o.party_name ILIKE %s
                  AND m.live_dead = 'LIVE'
                LIMIT 100
            """, (pattern,))
            rows = cur.fetchall()

            if rows:
                a_fuzzy += 1
                for row in rows:
                    cur.execute("""
                        INSERT INTO path_a_results
                            (plaintiff_clean, match_type, serial_number, mark_identification,
                             owner_name, status_code, live_dead)
                        VALUES (%s, 'fuzzy', %s, %s, %s, %s, %s)
                    """, (plaintiff, row[0], row[1], row[2], row[3], row[4]))
                cur.execute("UPDATE unique_plaintiffs SET query_status = 'fuzzy' WHERE id = %s", (pid,))
            else:
                a_miss += 1
                cur.execute("UPDATE unique_plaintiffs SET query_status = 'miss' WHERE id = %s", (pid,))

        conn.commit()
    print(f"  路径A完成: exact={a_exact}, fuzzy={a_fuzzy}, miss={a_miss}")

    # ---- 路径B: 品牌名查询 ----
    cur.execute("SELECT id, brand_clean FROM unique_brands ORDER BY id")
    brands = cur.fetchall()

    # 第1轮: 精确匹配
    print(f"\n  路径B 第1轮(精确): {len(brands)} 个品牌...")
    b_exact = 0
    b_miss_ids = []

    for i, (bid, brand) in enumerate(brands):
        if i % 500 == 0 and i > 0:
            print(f"    进度: {i}/{len(brands)}")
            conn.commit()

        cur.execute("""
            SELECT DISTINCT t.serial_number, t.mark_identification, o.party_name,
                   t.status_code, m.live_dead
            FROM trademarks t
            JOIN trademark_owners o ON t.serial_number = o.serial_number
            JOIN status_code_mapping m ON t.status_code = m.status_code
            WHERE t.mark_identification ILIKE %s
              AND m.live_dead = 'LIVE'
        """, (f'%{brand}%',))
        rows = cur.fetchall()

        if rows:
            b_exact += 1
            for row in rows:
                cur.execute("""
                    INSERT INTO path_b_results
                        (brand_clean, match_type, serial_number, mark_identification,
                         owner_name, status_code, live_dead)
                    VALUES (%s, 'exact', %s, %s, %s, %s, %s)
                """, (brand, row[0], row[1], row[2], row[3], row[4]))
            cur.execute("UPDATE unique_brands SET query_status = 'exact' WHERE id = %s", (bid,))
        else:
            b_miss_ids.append((bid, brand))

    conn.commit()
    print(f"  路径B 第1轮完成: exact={b_exact}, 待模糊={len(b_miss_ids)}")

    # 第2轮: 模糊匹配
    b_fuzzy = 0
    b_miss = 0
    if b_miss_ids:
        print(f"  路径B 第2轮(模糊): {len(b_miss_ids)} 个...")
        for i, (bid, brand) in enumerate(b_miss_ids):
            if i % 100 == 0 and i > 0:
                print(f"    进度: {i}/{len(b_miss_ids)}")
                conn.commit()

            words = brand.split()
            if len(words) >= 2:
                pattern = f'%{words[0]}%{words[1]}%'
            else:
                pattern = f'%{brand}%'

            cur.execute("""
                SELECT DISTINCT t.serial_number, t.mark_identification, o.party_name,
                       t.status_code, m.live_dead
                FROM trademarks t
                JOIN trademark_owners o ON t.serial_number = o.serial_number
                JOIN status_code_mapping m ON t.status_code = m.status_code
                WHERE t.mark_identification ILIKE %s
                  AND m.live_dead = 'LIVE'
                LIMIT 100
            """, (pattern,))
            rows = cur.fetchall()

            if rows:
                b_fuzzy += 1
                for row in rows:
                    cur.execute("""
                        INSERT INTO path_b_results
                            (brand_clean, match_type, serial_number, mark_identification,
                             owner_name, status_code, live_dead)
                        VALUES (%s, 'fuzzy', %s, %s, %s, %s, %s)
                    """, (brand, row[0], row[1], row[2], row[3], row[4]))
                cur.execute("UPDATE unique_brands SET query_status = 'fuzzy' WHERE id = %s", (bid,))
            else:
                b_miss += 1
                cur.execute("UPDATE unique_brands SET query_status = 'miss' WHERE id = %s", (bid,))

        conn.commit()
    print(f"  路径B完成: exact={b_exact}, fuzzy={b_fuzzy}, miss={b_miss}")

    cur.close()
    conn.close()


# =====================================================
# Step 3: 汇合判定
# =====================================================

def step3_merge():
    print("\n" + "=" * 60)
    print("Step 3: 汇合判定")
    print("=" * 60)

    conn = psycopg2.connect(DB_CONN)
    cur = conn.cursor()

    cur.execute("TRUNCATE matched_companies RESTART IDENTITY")

    # 获取所有案件
    cur.execute("""
        SELECT case_number, plaintiff_clean, brand_clean, brand_eq_plaintiff, nature_of_suit
        FROM tro_cases
    """)
    cases = cur.fetchall()

    stats = defaultdict(int)

    for case_num, plaintiff, brand, brand_eq, nature in cases:
        # 路径A结果: 该原告名下有哪些 owner
        # 空原告跳过路径A
        companies_a = set()
        if plaintiff:
            cur.execute("""
                SELECT DISTINCT owner_name FROM path_a_results WHERE plaintiff_clean = %s
            """, (plaintiff,))
            companies_a = set(r[0] for r in cur.fetchall())

        a_status = 'hit' if companies_a else 'miss'

        # 路径B结果: 该品牌归属哪些公司
        companies_b = set()
        b_status = 'skip'
        if not brand_eq and brand:
            cur.execute("""
                SELECT DISTINCT owner_name FROM path_b_results WHERE brand_clean = %s
            """, (brand,))
            companies_b = set(r[0] for r in cur.fetchall())
            b_status = 'hit' if companies_b else 'miss'

        # 汇合判定 (与 step2_v2.py run_step3 对齐)
        if a_status == 'hit' and b_status == 'hit':
            overlap = companies_a & companies_b
            if overlap:
                real_company = max(overlap, key=lambda c: validate_match(plaintiff, c))
                quality = 'exact'
                needs_review = False
                reason = None
            else:
                best_a = max(companies_a, key=lambda c: validate_match(plaintiff, c))
                score = validate_match(plaintiff, best_a)
                real_company = best_a
                quality = 'a_preferred'
                needs_review = score < 0.3
                reason = f'low_similarity: {score:.3f}' if needs_review else None
        elif a_status == 'hit' and b_status in ('miss', 'skip'):
            best_a = max(companies_a, key=lambda c: validate_match(plaintiff, c))
            score = validate_match(plaintiff, best_a)
            real_company = best_a
            quality = 'a_only'
            needs_review = score < 0.3
            reason = f'low_similarity: {score:.3f}' if needs_review else None
        elif a_status == 'miss' and b_status == 'hit':
            if len(companies_b) == 1:
                real_company = next(iter(companies_b))
                quality = 'b_only'
                needs_review = False
                reason = None
            else:
                best_b = max(companies_b, key=lambda c: validate_match(plaintiff, c))
                score = validate_match(plaintiff, best_b)
                real_company = best_b
                quality = 'b_only'
                needs_review = True
                reason = f'multiple_b_owners: {len(companies_b)}, best_score={score:.3f}'
        else:
            real_company = None
            quality = 'not_found'
            needs_review = True
            reason = f'nature={nature}'

        stats[quality] += 1

        cur.execute("""
            INSERT INTO matched_companies
                (case_number, plaintiff_clean, brand_clean, real_company, match_quality, needs_review, review_reason)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (case_num, plaintiff, brand, real_company, quality, needs_review, reason))

    conn.commit()

    print(f"  匹配结果分布:")
    for k, v in sorted(stats.items(), key=lambda x: -x[1]):
        print(f"    {k}: {v}")

    cur.execute("SELECT COUNT(*) FROM matched_companies WHERE needs_review")
    review_count = cur.fetchone()[0]
    print(f"  需人工审核: {review_count}")

    cur.close()
    conn.close()


# =====================================================
# Step 4: 品牌收集
# =====================================================

def step4_collect_brands():
    print("\n" + "=" * 60)
    print("Step 4: 品牌收集")
    print("=" * 60)

    conn = psycopg2.connect(DB_CONN)
    cur = conn.cursor()

    cur.execute("TRUNCATE company_brand_details RESTART IDENTITY")

    # 获取所有已确定的真实公司（去重）
    cur.execute("""
        SELECT DISTINCT real_company FROM matched_companies
        WHERE real_company IS NOT NULL
    """)
    companies = [r[0] for r in cur.fetchall()]
    print(f"  需查询公司数: {len(companies)}")

    total_brands = 0

    for i, company in enumerate(companies):
        if i % 200 == 0 and i > 0:
            print(f"    进度: {i}/{len(companies)}")
            conn.commit()

        cur.execute("""
            INSERT INTO company_brand_details
                (real_company, mark_identification, serial_number, registration_number,
                 status_code, live_dead, category, international_code,
                 international_desc_cn, international_desc_en, goods_services, first_use_date)
            SELECT DISTINCT
                %s,
                t.mark_identification,
                t.serial_number,
                t.registration_number,
                t.status_code,
                m.live_dead,
                m.category,
                tc.international_code,
                nc.name_cn,
                nc.name_en,
                ts.text,
                tc.first_use_date
            FROM trademarks t
            JOIN trademark_owners o ON t.serial_number = o.serial_number
            JOIN status_code_mapping m ON t.status_code = m.status_code
            LEFT JOIN trademark_classes tc ON t.serial_number = tc.serial_number
            LEFT JOIN nice_classification nc ON tc.international_code = nc.code
            LEFT JOIN trademark_statements ts ON t.serial_number = ts.serial_number
                AND ts.type_code LIKE 'GS%%'
            WHERE o.party_name = %s
              AND m.live_dead = 'LIVE'
        """, (company, company))

        total_brands += cur.rowcount

    conn.commit()

    cur.execute("SELECT COUNT(DISTINCT real_company) FROM company_brand_details")
    co_count = cur.fetchone()[0]
    cur.execute("SELECT COUNT(DISTINCT mark_identification) FROM company_brand_details")
    brand_count = cur.fetchone()[0]

    print(f"  收集完成: {co_count} 家公司, {brand_count} 个唯一品牌, {total_brands} 条品牌+类目记录")

    cur.close()
    conn.close()


# =====================================================
# Step 5: 输出 Excel
# =====================================================

def step5_output():
    print("\n" + "=" * 60)
    print("Step 5: 输出 Excel")
    print("=" * 60)

    conn = psycopg2.connect(DB_CONN)
    cur = conn.cursor()

    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def write_header(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

    def write_cell(ws, row, col, value):
        cell = ws.cell(row=row, column=col, value=value)
        cell.border = thin_border
        return cell

    # ===== 手动填写列样式 =====
    manual_font = Font(bold=True, color='000000', size=11)
    manual_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')

    # ===== Sheet 1: 公司概览 =====
    ws1 = wb.active
    ws1.title = '公司概览'
    headers1 = ['公司名称', '涉案数量', '案件号', '匹配质量', 'TRO原告名', 'TRO品牌名',
                 '活跃品牌数', '需人工审核', '起诉类型',
                 '✎ 确认公司名', '✎ 确认品牌名', '✎ 备注']
    write_header(ws1, headers1)
    # 手动填写列用黄色标题
    for col in [10, 11, 12]:
        cell = ws1.cell(row=1, column=col)
        cell.fill = manual_fill
        cell.font = manual_font

    cur.execute("""
        SELECT mc.real_company,
               COUNT(DISTINCT mc.case_number) as case_count,
               STRING_AGG(DISTINCT mc.case_number, ', ') as case_numbers,
               (ARRAY_AGG(mc.match_quality ORDER BY
                   CASE mc.match_quality
                       WHEN 'exact' THEN 1
                       WHEN 'a_preferred' THEN 2
                       WHEN 'a_only' THEN 3
                       WHEN 'ai_high' THEN 4
                       WHEN 'ai_medium' THEN 5
                       WHEN 'b_only' THEN 6
                       WHEN 'ai_low' THEN 7
                       WHEN 'b_uncertain' THEN 8
                       ELSE 9
                   END
               ))[1] as best_quality,
               STRING_AGG(DISTINCT mc.plaintiff_clean, '; ') as plaintiffs,
               STRING_AGG(DISTINCT mc.brand_clean, '; ') as brands,
               BOOL_OR(mc.needs_review) as needs_review,
               STRING_AGG(DISTINCT tc.nature_of_suit, ', ') as suit_types
        FROM matched_companies mc
        LEFT JOIN tro_cases tc ON mc.case_number = tc.case_number
        WHERE mc.real_company IS NOT NULL
        GROUP BY mc.real_company
        ORDER BY case_count DESC
    """)

    row = 2
    for r in cur.fetchall():
        company = r[0]
        # 查品牌数
        cur2 = conn.cursor()
        cur2.execute("""
            SELECT COUNT(DISTINCT mark_identification)
            FROM company_brand_details WHERE real_company = %s
            AND mark_identification IS NOT NULL AND mark_identification != ''
        """, (company,))
        brand_count = cur2.fetchone()[0]
        cur2.close()

        write_cell(ws1, row, 1, company)
        write_cell(ws1, row, 2, r[1])
        case_nums = r[2] or ''
        if len(case_nums) > 200:
            case_nums = case_nums[:200] + '...'
        write_cell(ws1, row, 3, case_nums)
        write_cell(ws1, row, 4, r[3])
        plaintiffs = r[4] or ''
        if len(plaintiffs) > 150:
            plaintiffs = plaintiffs[:150] + '...'
        write_cell(ws1, row, 5, plaintiffs)
        brands = r[5] or ''
        if len(brands) > 150:
            brands = brands[:150] + '...'
        write_cell(ws1, row, 6, brands)
        write_cell(ws1, row, 7, brand_count)
        cell = write_cell(ws1, row, 8, '是' if r[6] else '否')
        if r[6]:
            cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        write_cell(ws1, row, 9, r[7] or '')  # 起诉类型
        # 10-12列空白供人工填写
        row += 1

    overview_rows = row - 2
    widths1 = [45, 10, 50, 15, 40, 35, 12, 12, 15, 25, 25, 20]
    for i, w in enumerate(widths1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.auto_filter.ref = f'A1:L{row-1}'

    # ===== Sheet 2: 公司-品牌对照 =====
    ws2 = wb.create_sheet('公司-品牌对照')
    write_header(ws2, ['公司名称', '品牌名'])

    cur.execute("""
        SELECT DISTINCT real_company, mark_identification
        FROM company_brand_details
        WHERE mark_identification IS NOT NULL AND mark_identification != ''
        ORDER BY real_company, mark_identification
    """)

    row = 2
    for r in cur.fetchall():
        write_cell(ws2, row, 1, r[0])
        write_cell(ws2, row, 2, r[1])
        row += 1

    mapping_rows = row - 2
    ws2.column_dimensions['A'].width = 45
    ws2.column_dimensions['B'].width = 30
    ws2.auto_filter.ref = f'A1:B{row-1}'

    # ===== Sheet 3: 品牌明细 =====
    ws3 = wb.create_sheet('品牌明细')
    headers3 = ['公司名称', '品牌名', 'Serial Number', '注册号', '状态', 'Nice编号',
                 'Nice类目(中)', 'Nice类目(英)', '商品/服务描述']
    write_header(ws3, headers3)

    cur.execute("""
        SELECT DISTINCT real_company, mark_identification, serial_number, registration_number,
               live_dead || '/' || category,
               international_code, international_desc_cn, international_desc_en,
               goods_services
        FROM company_brand_details
        WHERE mark_identification IS NOT NULL AND mark_identification != ''
        ORDER BY real_company, mark_identification, international_code
    """)

    row = 2
    for r in cur.fetchall():
        write_cell(ws3, row, 1, r[0])
        write_cell(ws3, row, 2, r[1])
        write_cell(ws3, row, 3, r[2])
        reg = r[3] if r[3] and r[3] != '0000000' else '(申请中)'
        write_cell(ws3, row, 4, reg)
        write_cell(ws3, row, 5, r[4])
        write_cell(ws3, row, 6, r[5] or '')
        write_cell(ws3, row, 7, r[6] or '')
        write_cell(ws3, row, 8, r[7] or '')
        gs = r[8] or ''
        if len(gs) > 200:
            gs = gs[:200] + '...'
        write_cell(ws3, row, 9, gs)
        row += 1

    detail_rows = row - 2
    widths3 = [45, 25, 15, 15, 18, 10, 18, 20, 80]
    for i, w in enumerate(widths3, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.auto_filter.ref = f'A1:I{row-1}'

    # ===== Sheet 4: 未找到清单 =====
    ws4 = wb.create_sheet('未找到清单')
    headers4 = ['案件号', '原告名', '品牌名', '匹配质量', '起诉类型', '审核原因', '品牌=原告',
                 '✎ 确认公司名', '✎ 确认品牌名', '✎ 备注']
    write_header(ws4, headers4)
    for col in [8, 9, 10]:
        cell = ws4.cell(row=1, column=col)
        cell.fill = manual_fill
        cell.font = manual_font

    cur.execute("""
        SELECT mc.case_number, mc.plaintiff_clean, mc.brand_clean, mc.match_quality,
               tc.nature_of_suit, mc.review_reason, tc.brand_eq_plaintiff
        FROM matched_companies mc
        LEFT JOIN tro_cases tc ON mc.case_number = tc.case_number
        WHERE mc.match_quality = 'not_found'
        ORDER BY tc.nature_of_suit, mc.case_number
    """)

    row = 2
    for r in cur.fetchall():
        write_cell(ws4, row, 1, r[0])
        write_cell(ws4, row, 2, r[1] or '')
        write_cell(ws4, row, 3, r[2] or '')
        write_cell(ws4, row, 4, r[3])
        write_cell(ws4, row, 5, r[4] or '')
        write_cell(ws4, row, 6, r[5] or '')
        write_cell(ws4, row, 7, '是' if r[6] else '否')
        row += 1

    notfound_rows = row - 2
    widths4 = [20, 40, 30, 15, 15, 30, 10, 25, 25, 20]
    for i, w in enumerate(widths4, 1):
        ws4.column_dimensions[get_column_letter(i)].width = w
    ws4.auto_filter.ref = f'A1:J{row-1}'

    # ===== Sheet 5: 分类汇总 =====
    ws5 = wb.create_sheet('分类汇总')

    cur.execute("""
        SELECT tc.nature_of_suit,
               mc.match_quality,
               COUNT(*) as cnt
        FROM matched_companies mc
        LEFT JOIN tro_cases tc ON mc.case_number = tc.case_number
        GROUP BY tc.nature_of_suit, mc.match_quality
        ORDER BY tc.nature_of_suit, mc.match_quality
    """)
    cross_data = defaultdict(lambda: defaultdict(int))
    all_qualities = set()
    for suit, quality, cnt in cur.fetchall():
        suit = suit or '未知'
        cross_data[suit][quality] = cnt
        all_qualities.add(quality)

    quality_order = ['exact', 'a_preferred', 'a_only', 'ai_high', 'ai_medium', 'ai_low',
                     'b_only', 'b_uncertain', 'not_found']
    qualities = [q for q in quality_order if q in all_qualities]
    qualities += sorted(all_qualities - set(quality_order))

    headers5 = ['起诉类型'] + qualities + ['合计']
    write_header(ws5, headers5)

    row = 2
    for suit in sorted(cross_data.keys()):
        write_cell(ws5, row, 1, suit)
        total = 0
        for ci, q in enumerate(qualities, 2):
            v = cross_data[suit].get(q, 0)
            write_cell(ws5, row, ci, v)
            total += v
        write_cell(ws5, row, len(qualities) + 2, total)
        row += 1

    ws5.column_dimensions['A'].width = 15
    for i in range(2, len(qualities) + 3):
        ws5.column_dimensions[get_column_letter(i)].width = 14

    # ===== Sheet 6: 起诉类型汇总 =====
    ws6 = wb.create_sheet('起诉类型汇总')
    headers6 = ['起诉类型', '总案件数', '已匹配', '未找到', '匹配率']
    write_header(ws6, headers6)

    cur.execute("""
        SELECT tc.nature_of_suit,
               COUNT(*) as total,
               COUNT(*) FILTER (WHERE mc.match_quality != 'not_found') as matched,
               COUNT(*) FILTER (WHERE mc.match_quality = 'not_found') as not_found
        FROM matched_companies mc
        LEFT JOIN tro_cases tc ON mc.case_number = tc.case_number
        GROUP BY tc.nature_of_suit
        ORDER BY total DESC
    """)

    row = 2
    for suit, total, matched, not_found in cur.fetchall():
        write_cell(ws6, row, 1, suit or '未知')
        write_cell(ws6, row, 2, total)
        write_cell(ws6, row, 3, matched)
        write_cell(ws6, row, 4, not_found)
        rate = f"{matched/total*100:.1f}%" if total > 0 else "0%"
        write_cell(ws6, row, 5, rate)
        row += 1

    widths6 = [15, 12, 12, 12, 12]
    for i, w in enumerate(widths6, 1):
        ws6.column_dimensions[get_column_letter(i)].width = w

    wb.save(XLSX_OUTPUT)
    print(f"\n  已保存: {XLSX_OUTPUT}")
    print(f"  Sheet1 公司概览: {overview_rows} 行")
    print(f"  Sheet2 公司-品牌对照: {mapping_rows} 行")
    print(f"  Sheet3 品牌明细: {detail_rows} 行")
    print(f"  Sheet4 未找到清单: {notfound_rows} 行")
    print(f"  Sheet5 分类汇总")
    print(f"  Sheet6 起诉类型汇总")

    cur.close()
    conn.close()


# =====================================================
# Main
# =====================================================

if __name__ == '__main__':
    print("TRO → USPTO 批量匹配处理开始")
    print()

    step0_load_and_clean()
    step1_dedup()
    step2_batch_query()
    step3_merge()
    step4_collect_brands()
    step5_output()

    print("\n" + "=" * 60)
    print("全部完成!")
    print("=" * 60)
